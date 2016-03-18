#from pyexcel_ods import *
import openpyxl
import json
import time
import paramiko
#import xlrd
#data = get_data("/home/rupam.kumari/test_file.ods")
#print data,"dataa"
#from collections import OrderedDict
import os, subprocess
import shlex
book = openpyxl.load_workbook("/home/rupam.kumari/test_file.xlsx")
sheet = book.get_sheet_by_name('Sheet 1')
#print sheet.title,"sheet............"
#print sheet.nrows,"jhgghhhg"
offset = 1
rows = []
"""
for i in range(1,5):
	if i<= offset:
		continue
	l = []
	for j in range(1,5):
	
		#cell = sheet.cell(row= i,column =j)
		#v = cell.value
		l.append(sheet.cell(row= i,column =j).value)
	rows.append(l)
"""
for ws in book.worksheets:
	if ws
	for index,row in enumerate(ws.rows,start =2):
		print row[0].value

def ssh_connection(hostname, port,user,password):

	ssh=paramiko.SSHClient()
	ssh.load_system_host_keys()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	ssh.connect(hostname,port,user,password,timeout=5)
	ssh_shell = ssh.invoke_shell()
	ssh_shell.send('\n')
	time.sleep(1)
	for line in rows:
		cmd=shlex.split("ping %s -c 1 " % line[0])
		try:
   			output = subprocess.check_output(cmd)
		except subprocess.CalledProcessError,e:
   			print "The IP {0} is NotReacahble".format(cmd[-1])
			
			
		else:
   			print "The IP {0} is Reachable".format(cmd[-1])
			
		
 		
def main():
    host = "192.168.1.55"
    user = "teramatrix\\rupam.kumari"
    password = "ttpl@123"
    
    try:
        if host:
            ssh_connection(host, 22,user,password)
    except Exception,error:
	print error
        



if __name__=='__main__':
	main()

